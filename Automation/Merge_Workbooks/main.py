from openpyxl import load_workbook, Workbook

list_files = ['new_workbook', 'gastos']

# 1 - Create a new workbook
wb = Workbook()
name_file = 'merged_workbook.xlsx'

for nome in list_files:
    file = load_workbook(filename='../files/%s.xlsx' % nome)
    sheet = file[nome]
    max_rows = sheet.max_row
    max_columns = sheet.max_column
    ws = wb.create_sheet(title=nome)

    # 2 - Iterate over the rows and columns
    for i in range(1, max_rows+1):
        for j in range(1, max_columns+1):
            ws.cell(row=i, column=j, value=sheet.cell(row=i, column=j).value)

wb.remove(wb['Sheet'])
wb.save('../files/%s' % name_file)
