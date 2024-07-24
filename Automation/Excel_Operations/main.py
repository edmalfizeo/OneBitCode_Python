from openpyxl import load_workbook
from openpyxl.drawing.image import Image

wb = load_workbook('../Excel_txt/gastos.xlsx')
sheet = wb['Sheet']

total = 0
# 1 - Sum the values
for row in range(2, 9):
    total += int(sheet.cell(row, 2).value)

sheet['A9'] = 'Total'
sheet['B9'] = str(total)


# 2 - Merge cells
sheet['A10'] = 'Teste'
sheet.merge_cells('A10:B10')
sheet.unmerge_cells('A10:B10')
wb.save('../Excel_txt/gastos.xlsx')

# 3 - Insert image
img = Image('../Actions/bb.png')
sheet.add_image(img, 'A11')

# 4 - Delete cell
sheet.delete_rows(1)
sheet.delete_cols(3)
wb.save('../Excel_txt/gastos2.xlsx')







