from openpyxl import load_workbook

# 1 - Reading workbook and Selecting a sheet
wb = load_workbook(filename='../Excel/new_workbook.xlsx')
sheet1 = wb['Sheet1']

# 2 - Reading data from the workbook
for row in sheet1.iter_rows(values_only=True):
    print(row)