from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference, series

dict_year = {}

# 1 - Importing despesas
wb = load_workbook(filename='../files/Ano_e_Despesa.xlsx')
ws1 = wb['Despesa']
max_row = ws1.max_row

for i in range(2, max_row + 1):
    dict_year[ws1.cell(row=i, column=1).value] = {'Despesa': ws1.cell(row=i, column=2).value, 'Receita': 0}

# 2 - Importing receitas
wb = load_workbook(filename='../files/Ano_e_Receita.xlsx')
ws2 = wb['Receita']
max_row = ws2.max_row

for i in range(2, max_row + 1):
    dict_year[ws2.cell(row=i, column=1).value]['Receita'] = ws2.cell(row=i, column=2).value

# 3 - Creating new workbook
wb = Workbook()
ws = wb.active
ws.title = 'Despesa e Receita'

ws['A1'] = 'Ano'
ws['B1'] = 'Despesas'
ws['C1'] = 'Receitas'

i = 2
for key, value in dict_year.items():
    ws['A%s' % i] = key
    ws['B%s' % i] = value['Despesa']
    ws['C%s' % i] = value['Receita']
    i += 1

chart1 = BarChart()
chart1.type = 'col'
chart1.style = 12
chart1.title = 'Receitas x Despesas por Ano'
chart1.x_axis.title = 'Ano'
chart1.y_axis.title = 'R$'

data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=i)

year = Reference(ws, min_col=2, min_row=2, max_row=i)

chart1.add_data(data, titles_from_data=True)
chart1.set_categories(year)
chart1.shape = 4
ws.add_chart(chart1, 'A%s' % (i + 2))

wb.save('../files/Despesa_e_Receita.xlsx')
